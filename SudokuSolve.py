import openpyxl, sys
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook('SudoBook.xlsx')
type(wb)
check = True
count = 0
sheets = wb['Sheet1']
boxes = ['','','','','','','','','']
rows = ['','','','','','','','','']
columns = ['','','','','','','','','',]
for i in range(0,9):
    boxes[i] = '123456789'
    rows[i] = '123456789'
    columns[i] = '123456789'
marks = []              #Array to store possible values of each box
for i in range(0,9):     #Nested loop to create full possibilities for each box
    marks.append([])
    for j in range(0,9):
        marks[i].append('123456789')
def remove(a,b,c):              #Function: remove number c from a,b possibilities
    marks[a][b] = marks[a][b].replace(str(sheets[c].value),'')
def resolve(d):         #Function: erase marks from same column, row and box as given number
    change = True
    x = column_index_from_string(d[0])-1 #converts xcel x coord to marks x coord
    xstart = x//3*3         #y conversion
    y = int(d[1])-1
    ystart = y//3*3
    boxes[(y//3)*3+x//3] = boxes[(y//3)*3+x//3].replace(str(sheets[coords].value),'')
    rows[y] = rows[y].replace(str(sheets[coords].value),'')
    columns[x] = columns[x].replace(str(sheets[coords].value),'') 
    marks[x][y] = ''    #Erases all marks in box with known value
    for i in range(0,9):
        remove(i,y,d)   #Row
        remove(x,i,d)   #Column
        remove(xstart+i%3,ystart+i//3,d) #Box
for i in range(0,9):             #Erases marks based on starting numbers
    for j in range(0,9):
        coords = get_column_letter(j+1) + str(i+1)
        if sheets[coords].value != None:    #If box is not empty
            resolve(coords)                     #Erase the marks accordingly
change = True
while change == True:       #Continue working if things have changed
    change = False          #Stores whether something changes
    for h in range(0,9):
        for j in range(0,9):   
            if len(marks[j%3+h%3*3][j//3+h//3*3]) == 1: #If slot has only one mark:
                sheets[get_column_letter(j%3+1+h%3*3)+str(j//3+1+h//3*3)] = marks[j%3+h%3*3][j//3+h//3*3]   #fill in the slot
                resolve(get_column_letter(j%3+1+h%3*3)+str(j//3+1+h//3*3))
                change = True
        for char in boxes[h]:           #Find numbers with only one possible home
            loc = ''            #Stores possible homes for the number char
            for k in range(0,9):
                if char in marks[k%3+h%3*3][k//3+h//3*3]:
                    loc += str(k)
            if len(loc) == 1:   #If only one possible home for the number char
                sheets[get_column_letter(int(loc)%3+1+h%3*3)+str(int(loc)//3+1+h//3*3)] = char
                resolve(get_column_letter(int(loc)%3+1+h%3*3)+str(int(loc)//3+1+h//3*3))
                change = True
            elif len(loc) != 0 and int(loc[0])//3 == int(loc[len(loc)-1])//3 and change == False:     #Finds two marks alone in the same box horizontally
                #count += 1
                #if count == 10:
                    #print(char)
                for p in range(0,6):
                    x = ((int(loc[0])%3+h%3*3)//3*3+3+p)%9
                    y = int(loc[0])//3+h//3*3
                    marks[x][y] = marks[x][y].replace(char,'')
                    #if count == 10:
                        #check = False
                        #print(x)
                #sys.exit()
                

        for char in rows[h]:        #looks for sole place for number in row
            loc = ''
            for m in range(0,9):
                if char in marks[m][h]:
                    loc += str(m)
            if len(loc) == 1:
                sheets[get_column_letter(int(loc)+1)+str(h+1)] = char
                resolve(get_column_letter(int(loc)+1)+str(h+1))
                change = True
        for char in columns[h]:     #looks for sole place for number in column
            loc = ''
            for n in range(0,9):
                if char in marks[h][n]:
                    loc += str(n)
            if len(loc) == 1:
                sheets[get_column_letter(h+1)+str(int(loc)+1)] = char
                resolve(get_column_letter(h+1)+str(int(loc)+1))
                change = True

wb.save('SudoBook.xlsx')
